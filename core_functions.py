import win32gui     # (필수) 창 핸들 및 좌표 획득
import win32api
import win32con     # (필수) 창 상태 확인
import os
import sys
import tempfile
import shutil
import zipfile
import glob
import time 
import pythoncom
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill  
from openpyxl.utils import get_column_letter 
from PIL import Image, ImageGrab, ImageChops
from pynput.keyboard import Key, Controller 
from win32com.client import Dispatch, GetActiveObject
import hashlib
from common import _get_file_hash, capture_active_window, _clear_system_clipboard


try:
    import win32com.client
    import win32gui
    import win32con
except ImportError:
    print("pywin32 라이브러리가 필요합니다. pip install pywin32")
    sys.exit(1)

# --- 1. Scan Directory ---

def scan_directory(target_dir, output_dir):
    """
    지정된 디렉터리를 스캔하여 엑셀 파일로 저장하고,
    Output Dir에 동일한 구조의 빈 폴더/파일을 생성
    """
    
    target_dir_basename = os.path.basename(os.path.normpath(target_dir))
    output_excel_file = os.path.join(output_dir, f"{target_dir_basename}.xlsx")
    
    # 미러링 기본 경로 (Output Dir 하위에 원본 폴더명으로 생성)
    # 예: Output/TargetDirName/
    mirror_base_dir = os.path.join(output_dir, target_dir_basename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = target_dir_basename
    
    base_depth = target_dir.count(os.sep)
    file_cells_coords = [] 
    
    print(f"[DEBUG] 스캔 시작: {target_dir} -> {output_dir}")

    for root, dirs, files in os.walk(target_dir, topdown=True):
        
        # --- 1. Excel 생성 로직  ---
        current_depth = root.count(os.sep) - base_depth
        folder_name = "📁 " + os.path.basename(root)
        row = [None] * current_depth + [folder_name]
        
        if files:
            files_str = "\n".join(["┣ " + f for f in files])
            row.append(files_str)
        ws.append(row)
        
        if files:
            current_row_index = ws.max_row
            ws.row_dimensions[current_row_index].height = 13 * len(files)
            file_col_letter = chr(ord('A') + current_depth + 1)
            file_cells_coords.append(f"{file_col_letter}{current_row_index}")

        relative_path = os.path.relpath(root, target_dir)
        
        if relative_path == '.':
            dest_dir = mirror_base_dir
        else:
            dest_dir = os.path.join(mirror_base_dir, relative_path)
            
        os.makedirs(dest_dir, exist_ok=True)
        
        for f_name in files:
            dest_file_path = os.path.join(dest_dir, f_name)
            try:
                with open(dest_file_path, 'w') as f_empty:
                    pass
            except OSError as e:
                print(f"[WARN] 빈 파일 생성 실패: {dest_file_path}")

    # --- 열 너비 자동 조절 ---
    column_max_lengths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_idx = cell.column - 1 
                cell_value_str = str(cell.value)
                length = 0
                if "\n" in cell_value_str:
                    lines = cell_value_str.split('\n')
                    length = max(len(line) for line in lines)
                else:
                    length = len(cell_value_str)
                current_max = column_max_lengths.get(col_idx, 0)
                column_max_lengths[col_idx] = max(current_max, length)

    for col_idx, max_length in column_max_lengths.items():
        col_letter = get_column_letter(col_idx + 1) 
        ws.column_dimensions[col_letter].width = max_length + 2

    # --- 전체 셀 서식 적용 ---
    font_9pt = Font(size=9)
    align_top_no_wrap = Alignment(vertical='top', wrap_text=False)
    align_top_wrap = Alignment(vertical='top', wrap_text=True)

    gray_fill = PatternFill(start_color='BFBFBF',
                            end_color='BFBFBF',
                            fill_type='solid')

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font_9pt
            cell.alignment = align_top_no_wrap
            if cell.value is None:
                cell.fill = gray_fill
            
    for cell_coord in file_cells_coords:
        ws[cell_coord].alignment = align_top_wrap

    wb.save(output_excel_file)
    
    return f"디렉터리 스캔 완료!\n\n엑셀 파일: {output_excel_file}\n빈 파일 미러링: {mirror_base_dir}"

# --- 2. Convert To Image ---

def capture_ppt_slides(target_file, output_dir, base_filename, interval_sec):
    """
    ppt 파일을 열고 슬라이드를 한 페이지씩 이동하면서 화면을 캡처하고 파일로 저장
    """
    
    output_path = os.path.join(os.path.abspath(output_dir), base_filename)
    os.makedirs(output_path, exist_ok=True)
    
    powerpoint = None
    presentation = None

    try:
        print("[DEBUG] 1. PowerPoint Dispatch 및 Open 시도...")
        powerpoint = Dispatch("PowerPoint.Application")
        powerpoint.Visible = True
        file_path = os.path.abspath(target_file)

        presentation = powerpoint.Presentations.Open(file_path)
        slide_count = presentation.Slides.Count
        print(f"[DEBUG] 1. Open 성공. 총 슬라이드: {slide_count}개")

        # Powerpoint 윈도우 핸들 찾기 및 최대화/최상위 설정
        hwnd = win32gui.FindWindow("PPTFrameClass", None)
        if hwnd:
            win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(0.5)
        else:
            raise Exception("Powerpoint 윈도우 핸들을 찾을 수 없습니다. (클래스: PPTFrameClass)")

        for i in range(1, slide_count + 1):
            print("[DEBUG] 2. Slide-{i} 캡처 시도...") 
            slide = presentation.Slides(i)
            slide.Select()
            time.sleep(interval_sec)

            screenshot = capture_active_window(hwnd)
            output_file_path = os.path.join(output_path, f"slide_{i:03}.png")
            screenshot.save(output_file_path, "PNG")
            print("[DEBUG] 2. Slide-{i} 캡처 완료...") 

        print(f"[OK] {output_file_path} 저장 완료")

    except Exception as e:
        print(f"\n[!!!] 변환 작업 중 심각한 오류 발생: {e}\n")
        raise RuntimeError(f"PPT 변환 중 오류 발생: {e}")

    finally:
        if presentation:
            presentation.Close()
        if powerpoint:
            powerpoint.Quit()

    return f"PPT 슬라이드 {slide_count}개를 이미지로 저장 완료!\n{output_path}"

def capture_excel_sheets(target_file, output_dir, base_filename, interval_sec):

    """
    Excel 파일을 열고 각 시트의 내용을 화면 캡처하여 파일로 저장
    """
    output_path = os.path.join(os.path.abspath(output_dir), base_filename + "_Excel")
    os.makedirs(output_path, exist_ok=True)
    
    excel = None
    workbook = None
    sheet_count = 0

    try:
        print("[DEBUG] 1. Excel Dispatch 및 Open 시도...")
        excel = Dispatch("Excel.Application")
        excel.Visible = True
        file_path = os.path.abspath(target_file)

        workbook = excel.Workbooks.Open(file_path)
        sheet_count = workbook.Sheets.Count
        print(f"[DEBUG] 1. Open 성공. 총 시트: {sheet_count}개")

        # Excel 윈도우 핸들 찾기 및 최대화/최상위 설정
        # 엑셀의 클래스 이름은 보통 "XLMAIN"
        hwnd = win32gui.FindWindow("XLMAIN", None)
        if hwnd:
            win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(1.0)
        else:
            raise Exception("Excel 윈도우 핸들을 찾을 수 없습니다. (클래스: XLMAIN)")


        for i in range(1, sheet_count + 1):
            sheet = workbook.Sheets(i)
            sheet.Activate()
            time.sleep(interval_sec) 
            print(f"[DEBUG] 2. Sheet-{i} ('{sheet.Name}') 캡처 시도...") 

            # 화면 캡처
            screenshot = capture_active_window(hwnd)
            output_file_path = os.path.join(output_path, f"sheet_{i:03}_{sheet.Name.replace(' ', '_')}.png")
            screenshot.save(output_file_path, "PNG")
            print(f"[DEBUG] 2. Sheet-{i} 캡처 완료...") 

        print(f"[OK] {sheet_count}개 시트 이미지 저장 완료: {output_path}")

    except Exception as e:
        print(f"\n[!!!] Excel 변환 작업 중 심각한 오류 발생: {e}\n")
        raise RuntimeError(f"Excel 변환 중 오류 발생: {e}")

    finally:
        if workbook:
            workbook.Close(False) # 저장하지 않고 닫기
        if excel:
            excel.Quit()

    return f"Excel 시트 {sheet_count}개를 이미지로 저장 완료!\n{output_path}"

def capture_word_document(target_file, output_dir, base_filename, interval_sec):
    """
    Word 파일을 '한 페이지' 보기로 열고,
    'COM API(GoTo)' + '파일 해시 비교'로 모든 페이지를 캡처 (pynput 제거)
    """
    output_path = os.path.join(os.path.abspath(output_dir), base_filename + "_Word")
    os.makedirs(output_path, exist_ok=True)
    
    pythoncom.CoInitialize()
    word = None
    document = None
    page_count = 0
    prev_file_hash = None

    # Word VBA 상수 정의
    wdGoToPage = 1
    wdGoToNext = 2
    wdPrintView = 3          # '인쇄 모양' 보기
    wdRevisionsViewFinal = 0 # '최종본' 보기 (변경 내용/메모 숨기기)
    wdWindowStateMaximize = 1  # 창 최대화 상수

    try:
        print("[DEBUG] 1. Word Dispatch 및 Open 시도...")
        word = Dispatch("Word.Application")
        word.Visible = True
        file_path = os.path.abspath(target_file)
        document = word.Documents.Open(file_path)
        print("[DEBUG] 1. Open 성공.")

        print("[DEBUG] 1b. Word 윈도우 핸들('OpusApp') 탐색 시작...")

        hwnd = win32gui.FindWindow("OpusApp", None)
        if hwnd:
            print(f"[DEBUG] 1c. 윈도우 핸들 탐색 성공: {hwnd}")

            # 1. (복원) 최소화 상태일 수 있으므로 '복원'
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)
            word.Application.WindowState = wdWindowStateMaximize
            """
            # 2. (위치 강제) 크기 변경 없이 (0,0)으로 '이동'
            flags = win32con.SWP_SHOWWINDOW | win32con.SWP_NOSIZE
            win32gui.SetWindowPos(hwnd, -1, 0, 0, 0, 0, flags) 
            time.sleep(0.5) # 위치 이동 대기

            # 3. (최대화) COM 속성으로 최대화 *요청*
            word.Application.WindowState = wdWindowStateMaximize
            """
            # 4. (대기) *[중요]* Word가 최대화를 '완료'할 시간을 줍니다.
            print("[DEBUG] 1d. Word 창 최대화 대기 (1.5초)...")
            time.sleep(1.5)
            
            # 5. (포커스) *최대화가 완료된 후* 포커스를 설정합니다.
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(0.5)
            
            rect = win32gui.GetWindowRect(hwnd)
            print(f"[DEBUG] 1e. 창 최대화 및 포커스 완료. 현재 좌표: {rect}")
            # --- [수정 끝] ---

        else:
            raise Exception("Word 윈도우 핸들('OpusApp')을 찾지 못했습니다.")

        # --- [핵심 수정 2: 보기 모드 설정을 최대화 *이후*에 실행] ---
        try:
            # *[중요]* 이 작업은 창이 '완전히' 최대화된 후에 실행되어야 합니다.
            print("[DEBUG] 2. '인쇄 모양' 및 '한 페이지' 보기 모드로 변경 시도...")
            word.ActiveWindow.View.Type = wdPrintView 
            time.sleep(0.5) 
            word.ActiveWindow.View.RevisionsView = wdRevisionsViewFinal
            time.sleep(0.5) 
            
            # '한 페이지' 보기를 '최대화된 창' 크기에 맞춥니다.
            word.ActiveWindow.View.Zoom.PageFit = 1 
            print("[DEBUG] 2. 보기 모드 변경 성공.")
        except Exception as e:
            print(f"[WARN] 보기 모드 변경 실패 (오류: {e})")
        # --- [수정 끝] ---
            
        print("[DEBUG] Word 페이지 캡처 루프 시작 (파일 해시 비교 방식)...")        
        for i in range(1, 501): # 최대 500페이지
            
            print(f"[DEBUG] Word Page-{i} 캡처 시도...")
            try:
                # 캡처 직전 포커스 재확보
                win32gui.SetForegroundWindow(hwnd)
                time.sleep(interval_sec) 

                screenshot = capture_active_window(hwnd)
                print(f"[DEBUG] Window Handle = {hwnd}")
            except Exception as capture_err:
                print(f"[WARN] 캡처 실패(오류: {capture_err}). 루프를 중단합니다.")
                break

            output_file_path = os.path.join(output_path, f"{base_filename}_page_{i:03}.png")
            screenshot.save(output_file_path, "PNG")
            
            current_file_hash = _get_file_hash(output_file_path)
            
            print(f"[DEBUG] Page-{i} 비교: PrevHash={prev_file_hash}, CurrHash={current_file_hash}")

            if i > 1 and prev_file_hash == current_file_hash:
                print(f"[DEBUG] Page-{i}가 이전 페이지와 파일 해시가 동일하여 캡처를 중지합니다 (문서 끝).")
                try:
                    os.remove(output_file_path)
                    print(f"[DEBUG] 중복 저장된 {output_file_path} 파일을 삭제했습니다.")
                except Exception as e:
                    print(f"[WARN] 중복 파일 삭제 실패: {e}")
                break 
            
            prev_file_hash = current_file_hash
            page_count += 1
            print(f"[DEBUG] Word Page-{i} 캡처 및 저장 완료.")
            
            print(f"[DEBUG] COM API로 다음 페이지 이동 시도 (GoTo Page Next)...")
            try:
                document.Application.Selection.GoTo(wdGoToPage, wdGoToNext) 
                time.sleep(2.0)
            except Exception as e:
                print(f"[DEBUG] COM API 페이지 이동 실패 (문서 끝 추정: {e}). 루프를 중단합니다.")
                break
    except Exception as e:
        raise RuntimeError(f"Word 변환 중 오류 발생: {e}")

    finally:
        print("[DEBUG] 6. finally 블록 실행 (정리 시작)")
        if document:
            document.Close(False) 
        if word:
            word.Quit()
        pythoncom.CoUninitialize()

    return f"Word 문서 {page_count}페이지 이미지를 저장 완료!\n{output_path}"


def capture_pdf_document(target_file, output_dir, base_filename, interval_sec):
    """
    PDF 파일을 기본 뷰어로 열고, (포커스 + pynput)으로 PageDown을 전송하며
    '저장된 파일 해시'를 비교하여 모든 페이지를 캡처
    """
    output_path = os.path.join(os.path.abspath(output_dir), base_filename + "_PDF")
    os.makedirs(output_path, exist_ok=True)
    
    try:
        os.startfile(target_file)
    except Exception as e:
        raise RuntimeError(f"PDF 파일 열기 실패. 기본 뷰어 설정 확인: {e}")
    
    time.sleep(3.0) # 뷰어 로딩 대기

    hwnd = win32gui.FindWindow("AcrobatSDIWindow", None) # Adobe Acrobat
    if hwnd == 0:
        hwnd = win32gui.FindWindow("Chrome_WidgetWin_1", None) # Chrome/Edge
        print("[DEBUG] Adobe 뷰어를 찾지 못했습니다. Chrome/Edge 뷰어를 시도합니다.")
    if hwnd == 0:
        print("[DEBUG] 특정 뷰어를 찾을 수 없습니다. 현재 활성화된 창을 PDF 뷰어로 추정합니다.")
        hwnd = win32gui.GetForegroundWindow()
    if hwnd == 0:
        raise Exception("PDF 뷰어 창을 찾거나 활성화할 수 없습니다.")
    
    keyboard = Controller()
    page_count = 0
    
    # [수정] 이전 파일의 해시를 저장
    prev_file_hash = None
    
    try:
        win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)
        win32gui.SetForegroundWindow(hwnd)
        time.sleep(1.0)
        
        print("[DEBUG] PDF 페이지 캡처 루프 시작 (파일 해시 비교 방식)...")

        for i in range(1, 501): # 최대 500페이지
            
            print(f"[DEBUG] PDF Page-{i} 캡처 시도...")
            try:
                time.sleep(interval_sec)
                screenshot = capture_active_window(hwnd)
            except Exception as capture_err:
                print(f"[WARN] 캡처 실패(오류: {capture_err}). 루프를 중단합니다.")
                break
            
            # 1. 캡처한 이미지를 파일로 "먼저 저장"
            output_file_path = os.path.join(output_path, f"{base_filename}_page_{i:03}.png")
            screenshot.save(output_file_path, "PNG")
            
            # 2. 방금 저장된 파일의 해시 계산
            current_file_hash = _get_file_hash(output_file_path)
            
            print(f"[DEBUG] Page-{i} 비교: PrevHash={prev_file_hash}, CurrHash={current_file_hash}")

            # 3. 이전 파일 해시와 현재 파일 해시 비교
            if i > 1 and prev_file_hash == current_file_hash:
                # 두 파일 해시가 동일하면, PageDown이 안 먹힌 것 (문서 끝)
                print(f"[DEBUG] Page-{i}가 이전 페이지와 파일 해시가 동일하여 캡처를 중지합니다 (문서 끝).")
                
                # 마지막으로 저장된 중복 파일(page_i) 삭제
                try:
                    os.remove(output_file_path)
                    print(f"[DEBUG] 중복 저장된 {output_file_path} 파일을 삭제했습니다.")
                except Exception as e:
                    print(f"[WARN] 중복 파일 삭제 실패: {e}")
                
                break # 루프 중단
            
            # 4. (저장 성공) 현재 해시를 '이전 해시'로 저장하고 카운트 증가
            prev_file_hash = current_file_hash
            page_count += 1
            print(f"[DEBUG] PDF Page-{i} 캡처 및 저장 완료.")
            
            # 5. PageDown 키 전송
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(0.1) # 포커스 이동 대기
            print(f"[DEBUG] PageDown 키 전송 (pynput 방식)...")
            keyboard.press(Key.page_down)
            keyboard.release(Key.page_down)
            
        print("[DEBUG] 캡처 완료. 뷰어 창에 WM_CLOSE 메시지 전송...")
        win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
        time.sleep(3.0) 

        print(f"[OK] PDF 문서 {page_count}페이지 이미지 저장 완료: {output_path}")

    except Exception as e:
        print(f"\n[!!!] PDF 변환 작업 중 심각한 오류 발생: {e}\n")
        raise RuntimeError(f"PDF 변환 중 오류 발생: {e}")
    
    return f"PDF 문서 {page_count}페이지 이미지를 저장 완료!\n{output_path}"

def process_directory_for_images(target_dir, output_dir, interval_sec):
    """
    Target Dir 내의 모든 지원 파일을 검색하여 이미지 변환
    """
    target_dir = os.path.abspath(target_dir)
    output_dir = os.path.abspath(output_dir)

    if not os.path.exists(target_dir):
        raise FileNotFoundError(f"대상 폴더를 찾을 수 없습니다: {target_dir}")

    # 지원하는 확장자와 매핑되는 함수 정의
    conversion_map = {
        ".ppt": capture_ppt_slides,
        ".pptx": capture_ppt_slides,
        ".xls": capture_excel_sheets,
        ".xlsx": capture_excel_sheets,
        ".doc": capture_word_document,
        ".docx": capture_word_document,
        ".pdf": capture_pdf_document 
    }
    
    # 디렉터리 내 파일 검색
    all_files = os.listdir(target_dir)
    target_files = []
    
    for f in all_files:
        full_path = os.path.join(target_dir, f)
        if os.path.isfile(full_path):
            ext = os.path.splitext(f)[1].lower()
            if ext in conversion_map:
                target_files.append((full_path, ext))
    
    # 파일명 순으로 정렬 (옵션)
    target_files.sort()

    if not target_files:
        return f"지정된 폴더에 변환 가능한 파일이 없습니다.\n(대상: {target_dir})\n지원 확장자: ppt, xls, doc, pdf"

    success_count = 0
    fail_count = 0
    results_log = []

    print(f"\n[DEBUG] --- 배치 작업 시작 ---")
    print(f"[DEBUG] 총 {len(target_files)}개 변환 대상 파일 발견.")

    for i, (file_path, ext) in enumerate(target_files, 1):
        filename = os.path.basename(file_path)
        base_filename = os.path.splitext(filename)[0]
        
        print(f"\n>> [{i}/{len(target_files)}] 처리 중: {filename}")
        
        converter_func = conversion_map[ext]
        
        try:
            # 개별 변환 함수 호출
            # (각 함수는 output_dir 아래에 base_filename 폴더를 알아서 생성함)
            converter_func(file_path, output_dir, base_filename, interval_sec)
            
            success_count += 1
            results_log.append(f"[성공] {filename}")
            print(f">> [{i}/{len(target_files)}] 완료: {filename}")
            
        except Exception as e:
            fail_count += 1
            err_msg = f"[실패] {filename} : {str(e)}"
            print(err_msg)
            results_log.append(err_msg)
            
            # 오류 발생 시 잠시 대기 후 다음 파일 진행 (연속 오류 방지)
            time.sleep(2.0)

    # 최종 결과 리포트 생성
    summary = (
        f"작업이 완료되었습니다.\n\n"
        f"- 총 파일: {len(target_files)}개\n"
        f"- 성공: {success_count}개\n"
        f"- 실패: {fail_count}개\n\n"
        f"저장 경로: {output_dir}"
    )
    
    # 실패한 파일이 있다면 로그에 추가
    if fail_count > 0:
        summary += "\n\n[실패 목록]\n" + "\n".join([log for log in results_log if "[실패]" in log])
        
    return summary
    
# --- 3. Convert To PDF ---

def _numeric_sort_key(f):
    basename = os.path.splitext(os.path.basename(f))[0]
    try:
        # 파일명이 "slide_001.png" 같은 경우, "001"을 숫자로 변환하여 정렬
        # 숫자가 아닌 경우(예: "__MACOSX")는 basename으로 정렬
        return int(basename)
    except ValueError:
        return basename


def convert_to_pdf(target_root_dir, output_root_dir):
    """
    Target Root Dir 하위에 있는 '각 폴더'를 하나의 PDF로 변환
    """

    target_root_dir = os.path.abspath(target_root_dir)
    output_root_dir = os.path.abspath(output_root_dir)
    img_extensions = ('.png', '.jpg', '.jpeg', '.bmp', '.gif')

    # 하위 디렉터리 탐색
    sub_dirs = [
        d for d in os.listdir(target_root_dir) 
        if os.path.isdir(os.path.join(target_root_dir, d))
    ]
    
    if not sub_dirs:
        return f"지정된 Target Dir 내에 처리할 하위 폴더가 없습니다.\n({target_root_dir})"

    print(f"[DEBUG] PDF 변환 배치 시작. 대상 폴더: {len(sub_dirs)}개")
    
    success_count = 0
    fail_count = 0
    results_log = []

    for folder_name in sub_dirs:
        current_img_dir = os.path.join(target_root_dir, folder_name)
        
        # [수정] PDF 파일명은 폴더명과 동일하게 설정
        pdf_filename = f"{folder_name}.pdf"
        output_pdf_path = os.path.join(output_root_dir, pdf_filename)

        print(f"\n>> 처리 중: {folder_name} -> {pdf_filename}")

        try:
            # 1. 해당 폴더 내 이미지 파일 검색
            image_files = [
                os.path.join(current_img_dir, f) 
                for f in os.listdir(current_img_dir) 
                if os.path.splitext(f)[1].lower() in img_extensions
            ]

            if not image_files:
                print(f"[SKIP] '{folder_name}' 폴더에 이미지가 없어 건너뜁니다.")
                results_log.append(f"[SKIP] {folder_name} (이미지 없음)")
                continue

            # 2. 정렬
            image_files.sort(key=_numeric_sort_key)

            # 3. Pillow 이미지 로드 및 PDF 변환
            images_pil = []
            for img_path in image_files:
                try:
                    img = Image.open(img_path).convert('RGB')
                    images_pil.append(img)
                except Exception as img_err:
                    print(f"[WARN] 이미지 로드 실패 ({img_path}): {img_err}")

            if not images_pil:
                print(f"[SKIP] '{folder_name}' 폴더에서 유효한 이미지를 로드하지 못했습니다.")
                continue
            
            # 4. PDF 저장
            images_pil[0].save(
                output_pdf_path,
                save_all=True,
                append_images=images_pil[1:]
            )
            
            success_count += 1
            print(f"[OK] 저장 완료: {output_pdf_path}")
            results_log.append(f"[성공] {folder_name}.pdf")

        except Exception as e:
            fail_count += 1
            err_msg = f"[실패] {folder_name} : {str(e)}"
            print(err_msg)
            results_log.append(err_msg)

    # 최종 결과 리포트
    summary = (
        f"PDF 일괄 변환 완료!\n\n"
        f"- 총 폴더 스캔: {len(sub_dirs)}개\n"
        f"- 생성 성공: {success_count}개\n"
        f"- 실패: {fail_count}개\n\n"
        f"저장 경로: {output_root_dir}"
    )

    if fail_count > 0:
        summary += "\n\n[처리 로그]\n" + "\n".join(results_log)
        
    return summary

# --- 4. Remove DRM (Content Copy & Save) ---

def remove_drm_ppt(target_file, output_path):
    """
    PPT 파일을 열어 페이지 설정(크기)을 맞춘 후,
    슬라이드를 모두 복사하여 새 파일에 붙여넣어 저장
    """
    powerpoint = None
    source_pres = None
    new_pres = None
    
    try:
        powerpoint = Dispatch("PowerPoint.Application")
        powerpoint.Visible = True
        powerpoint.DisplayAlerts = 0 
        
        # 1. 원본 열기
        source_pres = powerpoint.Presentations.Open(os.path.abspath(target_file))
        
        # 2. 새 프레젠테이션 생성
        new_pres = powerpoint.Presentations.Add()
        
        # 원본의 슬라이드 크기(너비/높이)를 새 파일에 적용
        source_setup = source_pres.PageSetup
        new_setup = new_pres.PageSetup
        
        new_setup.SlideWidth = source_setup.SlideWidth
        new_setup.SlideHeight = source_setup.SlideHeight
        
        # 3. 슬라이드 복사 및 붙여넣기
        if source_pres.Slides.Count > 0:
            source_pres.Slides.Range().Copy()
            time.sleep(1.0) # 클립보드 안정화 대기
            new_pres.Slides.Paste()
        
        # 저장 전 원본 먼저 닫기
        source_pres.Close()
        source_pres = None 
        
        new_pres.SaveAs(os.path.abspath(output_path))
        print(f"[OK] PPT 저장 완료: {output_path}")
        
    except Exception as e:
        raise RuntimeError(f"PPT 처리 실패: {e}")
    finally:
        _clear_system_clipboard()
        if source_pres: 
            try: source_pres.Close()
            except: pass
        if new_pres: 
            try: new_pres.Close()
            except: pass
        if powerpoint: 
            try: powerpoint.Quit()
            except: pass

def remove_drm_excel(target_file, output_path):
    """Excel 파일을 열어 시트를 새 통합 문서로 복사하여 저장"""
    excel = None
    source_wb = None
    new_wb = None
    
    try:
        excel = Dispatch("Excel.Application")
        excel.Visible = True
        excel.DisplayAlerts = False 
        
        source_wb = excel.Workbooks.Open(os.path.abspath(target_file))
        
        # 시트 전체 복사 (인자 없이 Copy하면 새 워크북 생성됨)
        source_wb.Sheets.Copy()
        new_wb = excel.ActiveWorkbook
        
        # [수정 핵심] 저장하기 전에 원본 파일을 먼저 닫아야 "같은 이름으로 열려있음" 에러 방지
        source_wb.Close(False)
        source_wb = None 
        
        # 새 파일 저장
        new_wb.SaveAs(os.path.abspath(output_path))
        print(f"[OK] Excel 저장 완료: {output_path}")
        
    except Exception as e:
        raise RuntimeError(f"Excel 처리 실패: {e}")
    finally:
        # 클립보드 비우기
        _clear_system_clipboard()

        # 명시적 자원 해제 및 종료
        if source_wb: 
            try: source_wb.Close(False)
            except: pass
        if new_wb: 
            try: new_wb.Close(False)
            except: pass
        if excel: 
            excel.DisplayAlerts = True
            try: excel.Quit()
            except: pass

def remove_drm_word(target_file, output_path):
    """
    Word 파일을 열어 페이지 설정을 맞춘 후,
    전체 내용을 복사하여 새 문서에 붙여넣고 저장
    """
    word = None
    source_doc = None
    new_doc = None
    
    try:
        word = Dispatch("Word.Application")
        word.Visible = True
        
        source_doc = word.Documents.Open(os.path.abspath(target_file))
        
        # 새 문서 생성
        new_doc = word.Documents.Add()

        # [수정] 페이지 설정(PageSetup) 동기화 (첫 번째 섹션 기준)
        try:
            source_setup = source_doc.PageSetup
            new_setup = new_doc.PageSetup
            
            new_setup.Orientation = source_setup.Orientation  # 가로/세로 방향
            new_setup.PageWidth = source_setup.PageWidth      # 용지 너비
            new_setup.PageHeight = source_setup.PageHeight    # 용지 높이
            
            # 여백 설정 복사 (필요 시)
            new_setup.TopMargin = source_setup.TopMargin
            new_setup.BottomMargin = source_setup.BottomMargin
            new_setup.LeftMargin = source_setup.LeftMargin
            new_setup.RightMargin = source_setup.RightMargin
        except Exception as setup_err:
            print(f"[WARN] Word 페이지 설정 복사 중 일부 실패(무시): {setup_err}")
        
        # 내용 복사
        source_doc.Content.Copy() # WholeStory보다 Content.Copy가 안정적일 수 있음
        time.sleep(0.5)
        
        # 붙여넣기
        new_doc.Range().Paste()
        
        # 원본 닫기
        source_doc.Close(False)
        source_doc = None
        
        new_doc.SaveAs(os.path.abspath(output_path), FileFormat=16) # docx
        print(f"[OK] Word 저장 완료: {output_path}")
        
    except Exception as e:
        raise RuntimeError(f"Word 처리 실패: {e}")
    finally:
        _clear_system_clipboard()
        if source_doc: 
            try: source_doc.Close(False)
            except: pass
        if new_doc: 
            try: new_doc.Close(False)
            except: pass
        if word: 
            try: word.Quit()
            except: pass
            
def remove_drm_pdf_via_image(target_file, output_path):
    """
    PDF -> 이미지 캡처(기능2) -> PDF 병합(기능3) 방식을 사용하여 재생성
    """
    
    # 1. 임시 폴더 생성 (Output 폴더 내부에 hidden temp folder 생성)
    base_dir = os.path.dirname(output_path)
    filename = os.path.basename(output_path)
    filename_no_ext = os.path.splitext(filename)[0]
    
    # 임시 이미지 저장 경로: OutputDir/_temp_filename/
    temp_img_dir = os.path.join(base_dir, f"_temp_{filename_no_ext}")
    if os.path.exists(temp_img_dir):
        shutil.rmtree(temp_img_dir)
    os.makedirs(temp_img_dir)

    try:
        # 2. 이미지 캡처 (기존 capture_pdf_document 함수 재사용)
        # capture_pdf_document는 내부적으로 폴더를 하나 더 생성하므로 경로 조정 필요
        # capture_pdf_document(file, output_root, base_name, interval)
        print(f"[DEBUG] PDF 이미지 캡처 시작: {target_file}")
        
        # 캡처 속도(interval)는 0.5초로 설정 (필요 시 조정)
        capture_pdf_document(target_file, temp_img_dir, "capture", 0.5)
        
        # capture_pdf_document는 'temp_img_dir/capture_PDF' 폴더에 이미지를 저장함
        actual_img_dir = os.path.join(temp_img_dir, "capture_PDF")
        
        if not os.path.exists(actual_img_dir):
             raise RuntimeError("PDF 캡처 실패: 이미지 폴더가 생성되지 않았습니다.")

        # 3. 이미지들을 하나로 묶어 PDF 생성 (기존 convert_to_pdf 로직의 축소판)
        img_files = [
            os.path.join(actual_img_dir, f) 
            for f in os.listdir(actual_img_dir) 
            if f.lower().endswith('.png')
        ]
        
        if not img_files:
            raise RuntimeError("캡처된 이미지가 없습니다.")

        # 정렬 (page_001, page_002 ...)
        img_files.sort() 

        images_pil = []
        for img_p in img_files:
            try:
                img = Image.open(img_p).convert('RGB')
                images_pil.append(img)
            except Exception as e:
                print(f"[WARN] 이미지 로드 실패: {e}")

        if not images_pil:
            raise RuntimeError("변환할 유효한 이미지가 없습니다.")

        # 4. 최종 PDF 저장
        images_pil[0].save(
            output_path,
            save_all=True,
            append_images=images_pil[1:]
        )
        print(f"[OK] PDF 재생성 완료: {output_path}")

    except Exception as e:
        raise RuntimeError(f"PDF 이미지 변환 방식 실패: {e}")
        
    finally:
        # 5. 임시 폴더 삭제 (정리)
        if os.path.exists(temp_img_dir):
            try:
                shutil.rmtree(temp_img_dir)
                print(f"[DEBUG] 임시 폴더 삭제 완료: {temp_img_dir}")
            except Exception as e:
                print(f"[WARN] 임시 폴더 삭제 실패: {e}")


def process_remove_drm(target_dir, output_dir):
    """
    Target Dir 내의 파일을 읽어 DRM 제거 후 Output Dir에 '_해제' 접미사를 붙여 저장
    """
    target_dir = os.path.abspath(target_dir)
    output_dir = os.path.abspath(output_dir)

    drm_map = {
        ".ppt": remove_drm_ppt,
        ".pptx": remove_drm_ppt,
        ".xls": remove_drm_excel,
        ".xlsx": remove_drm_excel,
        ".doc": remove_drm_word,
        ".docx": remove_drm_word,
        ".pdf": remove_drm_pdf_via_image  
    }
    
    all_files = os.listdir(target_dir)
    target_files = []
    
    for f in all_files:
        full_path = os.path.join(target_dir, f)
        if os.path.isfile(full_path):
            ext = os.path.splitext(f)[1].lower()
            if ext in drm_map:
                target_files.append((full_path, ext))
                
    if not target_files:
        return f"변환할 파일이 없습니다.\n({target_dir})"

    success_count = 0
    fail_count = 0
    results_log = []
    
    print(f"[DEBUG] DRM 제거 배치 시작. 총 {len(target_files)}개")

    for i, (file_path, ext) in enumerate(target_files, 1):
        filename = os.path.basename(file_path)
        print(f"\n>> [{i}/{len(target_files)}] DRM 처리 중: {filename}")
        
        # 파일명 분리 및 '_해제' 접미사 추가
        base_name, file_ext = os.path.splitext(filename)
        new_filename = f"{base_name}_해제{file_ext}"
        
        output_file_path = os.path.join(output_dir, new_filename)        
        # 구버전 확장자(doc, xls, ppt)는 신규 포맷(x)으로 저장하는 것이 안정적임 (선택사항)
        if ext == '.ppt': output_file_path += 'x'
        elif ext == '.doc': output_file_path += 'x'
        elif ext == '.xls': output_file_path += 'x'

        func = drm_map[ext]
        
        try:
            func(file_path, output_file_path)
            success_count += 1
            results_log.append(f"[성공] {filename}")
        except Exception as e:
            fail_count += 1
            err_msg = f"[실패] {filename} : {str(e)}"
            print(err_msg)
            results_log.append(err_msg)
            time.sleep(1.0)

    summary = (
        f"DRM 제거(재저장) 작업 완료!\n\n"
        f"- 총 파일: {len(target_files)}개\n"
        f"- 성공: {success_count}개\n"
        f"- 실패: {fail_count}개\n\n"
        f"저장 경로: {output_dir}"
    )
    
    if fail_count > 0:
        summary += "\n\n[실패 목록]\n" + "\n".join([log for log in results_log if "[실패]" in log])
        
    return summary