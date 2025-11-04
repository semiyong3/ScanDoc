import win32gui     # (필수) 창 핸들 및 좌표 획득
import win32api
import win32con     # (필수) 창 상태 확인
import os
import sys
import tempfile
import shutil
import zipfile
import glob
import time 
import pythoncom
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill  
from openpyxl.utils import get_column_letter 
from PIL import Image, ImageGrab, ImageChops
from pynput.keyboard import Key, Controller 
from win32com.client import Dispatch
import hashlib
from common import _get_file_hash, capture_active_window


try:
    import win32com.client
    import win32gui
    import win32con
except ImportError:
    print("pywin32 라이브러리가 필요합니다. pip install pywin32")
    sys.exit(1)

# --- 1. Scan Directory ---

def scan_directory(target_dir, output_file):
    """
    지정된 디렉터리를 스캔하여 엑셀 파일로 저장 
    """
    wb = Workbook()
    ws = wb.active
    
    sheet_name = os.path.splitext(os.path.basename(output_file))[0]
    ws.title = sheet_name
    
    base_depth = target_dir.count(os.sep)
    file_cells_coords = [] 
    
    for root, dirs, files in os.walk(target_dir, topdown=True):
        current_depth = root.count(os.sep) - base_depth
        
        folder_name = "📁 " + os.path.basename(root)
        row = [None] * current_depth + [folder_name]
        
        if files:
            files_str = "\n".join(["┣ " + f for f in files])
            row.append(files_str)
            
        ws.append(row)
        
        if files:
            current_row_index = ws.max_row
            ws.row_dimensions[current_row_index].height = 13 * len(files)
            file_col_letter = chr(ord('A') + current_depth + 1)
            file_cells_coords.append(f"{file_col_letter}{current_row_index}")

    # --- 열 너비 자동 조절 ---
    column_max_lengths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_idx = cell.column - 1 
                cell_value_str = str(cell.value)
                length = 0
                if "\n" in cell_value_str:
                    lines = cell_value_str.split('\n')
                    length = max(len(line) for line in lines)
                else:
                    length = len(cell_value_str)
                current_max = column_max_lengths.get(col_idx, 0)
                column_max_lengths[col_idx] = max(current_max, length)

    for col_idx, max_length in column_max_lengths.items():
        col_letter = get_column_letter(col_idx + 1) 
        ws.column_dimensions[col_letter].width = max_length + 2

    # --- 전체 셀 서식 적용 ---
    
    font_9pt = Font(size=9)
    align_top_no_wrap = Alignment(vertical='top', wrap_text=False)
    align_top_wrap = Alignment(vertical='top', wrap_text=True)

    # 빈셀은 회색으로 채워서 가독성을 높임
    gray_fill = PatternFill(start_color='BFBFBF',
                            end_color='BFBFBF',
                            fill_type='solid')

    for row in ws.iter_rows():
        for cell in row:
            # 1. 기본 폰트 및 정렬 적용
            cell.font = font_9pt
            cell.alignment = align_top_no_wrap
            
            # 2. (신규) 값이 없는 셀(None)인 경우 회색으로 채우기
            if cell.value is None:
                cell.fill = gray_fill
            
    # 3. 파일 목록 셀에만 '줄바꿈 허용' 서식 덮어쓰기
    for cell_coord in file_cells_coords:
        ws[cell_coord].alignment = align_top_wrap

    wb.save(output_file)
    return f"디렉터리 스캔 완료!\n{output_file}"

# --- 2. Convert To Image ---

def capture_ppt_slides(target_file, output_dir, base_filename):
    """
    ppt 파일을 열고 슬라이드를 한 페이지씩 이동하면서 화면을 캡처하고 파일로 저장
    """
    
    output_path = os.path.join(os.path.abspath(output_dir), base_filename)
    os.makedirs(output_path, exist_ok=True)
    
    powerpoint = None
    presentation = None

    try:
        print("[DEBUG] 1. PowerPoint Dispatch 및 Open 시도...")
        powerpoint = Dispatch("PowerPoint.Application")
        powerpoint.Visible = True
        file_path = os.path.abspath(target_file)

        presentation = powerpoint.Presentations.Open(file_path)
        slide_count = presentation.Slides.Count
        print(f"[DEBUG] 1. Open 성공. 총 슬라이드: {slide_count}개")

        # Powerpoint 윈도우 핸들 찾기 및 최대화/최상위 설정
        hwnd = win32gui.FindWindow("PPTFrameClass", None)
        if hwnd:
            win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(0.5)
        else:
            raise Exception("Powerpoint 윈도우 핸들을 찾을 수 없습니다. (클래스: PPTFrameClass)")

        for i in range(1, slide_count + 1):
            print("[DEBUG] 2. Slide-{i} 캡처 시도...") 
            slide = presentation.Slides(i)
            slide.Select()
            time.sleep(0.5) 

            screenshot = capture_active_window(hwnd)
            output_file_path = os.path.join(output_path, f"slide_{i:03}.png")
            screenshot.save(output_file_path, "PNG")
            print("[DEBUG] 2. Slide-{i} 캡처 완료...") 

        print(f"[OK] {output_file_path} 저장 완료")

    except Exception as e:
        print(f"\n[!!!] 변환 작업 중 심각한 오류 발생: {e}\n")
        raise RuntimeError(f"PPT 변환 중 오류 발생: {e}")

    finally:
        if presentation:
            presentation.Close()
        if powerpoint:
            powerpoint.Quit()

    return f"PPT 슬라이드 {slide_count}개를 이미지로 저장 완료!\n{output_path}"

def capture_excel_sheets(target_file, output_dir, base_filename):

    """
    Excel 파일을 열고 각 시트의 내용을 화면 캡처하여 파일로 저장
    """
    output_path = os.path.join(os.path.abspath(output_dir), base_filename + "_Excel")
    os.makedirs(output_path, exist_ok=True)
    
    excel = None
    workbook = None
    sheet_count = 0

    try:
        print("[DEBUG] 1. Excel Dispatch 및 Open 시도...")
        excel = Dispatch("Excel.Application")
        excel.Visible = True
        file_path = os.path.abspath(target_file)

        workbook = excel.Workbooks.Open(file_path)
        sheet_count = workbook.Sheets.Count
        print(f"[DEBUG] 1. Open 성공. 총 시트: {sheet_count}개")

        # Excel 윈도우 핸들 찾기 및 최대화/최상위 설정
        # 엑셀의 클래스 이름은 보통 "XLMAIN"
        hwnd = win32gui.FindWindow("XLMAIN", None)
        if hwnd:
            win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(0.5)
        else:
            raise Exception("Excel 윈도우 핸들을 찾을 수 없습니다. (클래스: XLMAIN)")


        for i in range(1, sheet_count + 1):
            sheet = workbook.Sheets(i)
            sheet.Activate()
            time.sleep(0.5) 
            print(f"[DEBUG] 2. Sheet-{i} ('{sheet.Name}') 캡처 시도...") 

            # 화면 캡처
            screenshot = capture_active_window(hwnd)
            output_file_path = os.path.join(output_path, f"sheet_{i:03}_{sheet.Name.replace(' ', '_')}.png")
            screenshot.save(output_file_path, "PNG")
            print(f"[DEBUG] 2. Sheet-{i} 캡처 완료...") 

        print(f"[OK] {sheet_count}개 시트 이미지 저장 완료: {output_path}")

    except Exception as e:
        print(f"\n[!!!] Excel 변환 작업 중 심각한 오류 발생: {e}\n")
        raise RuntimeError(f"Excel 변환 중 오류 발생: {e}")

    finally:
        if workbook:
            workbook.Close(False) # 저장하지 않고 닫기
        if excel:
            excel.Quit()

    return f"Excel 시트 {sheet_count}개를 이미지로 저장 완료!\n{output_path}"

def capture_word_document(target_file, output_dir, base_filename):
    """
    Word 파일을 (DRM이 걸린) 임시 PDF로 변환한 후,
    capture_pdf_document 함수를 호출하여 캡처를 수행
    """
    # [수정] 1. COM 초기화
    pythoncom.CoInitialize()
    
    word = None
    document = None
    
    # [수정] 2. 임시 PDF 파일 경로 생성
    temp_pdf_path = os.path.join(tempfile.gettempdir(), f"{base_filename}_temp.pdf")
    print(f"[DEBUG] 임시 PDF 경로 설정: {temp_pdf_path}")
    
    try:
        # [수정] 3. Word를 열고 PDF로 '다른 이름으로 저장'
        print("[DEBUG] 1. Word Dispatch 및 Open 시도...")
        word = Dispatch("Word.Application")
        word.Visible = False # 백그라운드에서 실행
        file_path = os.path.abspath(target_file)
        document = word.Documents.Open(file_path)
        print("[DEBUG] 1. Open 성공.")

        print("[DEBUG] 2. PDF로 SaveAs 시도... (wdFormatPDF = 17)")
        # (VBA 상수 wdFormatPDF = 17)
        document.SaveAs(temp_pdf_path, FileFormat=17)
        print("[DEBUG] 2. 임시 PDF 저장 성공.")

    except Exception as e:
        print(f"\n[!!!] Word -> PDF 변환 중 심각한 오류 발생: {e}\n")
        raise RuntimeError(f"Word를 PDF로 자동 저장하는 중 오류 발생: {e}")
    
    finally:
        # [수정] 4. Word 프로세스 즉시 종료
        if document:
            document.Close(False) 
        if word:
            word.Quit()
        pythoncom.CoUninitialize()
        print("[DEBUG] 3. Word 종료 완료.")

    # --- [수정] 5. PDF 캡처 로직 호출 ---
    # 이제 임시 PDF 파일을 대상으로, 기존의 PDF 캡처 함수를 호출
    print(f"[DEBUG] 4. capture_pdf_document 함수 호출 (대상: {temp_pdf_path})")
    try:
        # (중요) PDF 캡처 함수가 "Word"가 아닌 "PDF" 폴더를 생성하도록 base_filename 수정
        pdf_base_filename = base_filename + "_Word"
        
        # 기존 PDF 캡처 함수 재사용
        result_msg = capture_pdf_document(temp_pdf_path, output_dir, pdf_base_filename)
        
    finally:
        # [수정] 6. PDF 캡처가 성공하든 실패하든, 임시 PDF 파일 삭제
        if os.path.exists(temp_pdf_path):
            try:
                os.remove(temp_pdf_path)
                print(f"[DEBUG] 5. 임시 PDF 파일 삭제 완료: {temp_pdf_path}")
            except Exception as e:
                print(f"[WARN] 임시 PDF 파일 삭제 실패: {e}")

    # PDF 캡처 함수의 결과 메시지를 반환
    return result_msg



def capture_pdf_document(target_file, output_dir, base_filename):
    """
    PDF 파일을 기본 뷰어로 열고, (포커스 + pynput)으로 PageDown을 전송하며
    '저장된 파일 해시'를 비교하여 모든 페이지를 캡처
    """
    output_path = os.path.join(os.path.abspath(output_dir), base_filename + "_PDF")
    os.makedirs(output_path, exist_ok=True)
    
    try:
        os.startfile(target_file)
    except Exception as e:
        raise RuntimeError(f"PDF 파일 열기 실패. 기본 뷰어 설정 확인: {e}")
    
    time.sleep(3.0) # 뷰어 로딩 대기

    hwnd = win32gui.FindWindow("AcrobatSDIWindow", None) # Adobe Acrobat
    if hwnd == 0:
        hwnd = win32gui.FindWindow("Chrome_WidgetWin_1", None) # Chrome/Edge
        print("[DEBUG] Adobe 뷰어를 찾지 못했습니다. Chrome/Edge 뷰어를 시도합니다.")
    if hwnd == 0:
        print("[DEBUG] 특정 뷰어를 찾을 수 없습니다. 현재 활성화된 창을 PDF 뷰어로 추정합니다.")
        hwnd = win32gui.GetForegroundWindow()
    if hwnd == 0:
        raise Exception("PDF 뷰어 창을 찾거나 활성화할 수 없습니다.")
    
    keyboard = Controller()
    page_count = 0
    
    # [수정] 이전 파일의 해시를 저장
    prev_file_hash = None
    
    try:
        win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)
        win32gui.SetForegroundWindow(hwnd)
        time.sleep(1.0)
        
        print("[DEBUG] PDF 페이지 캡처 루프 시작 (파일 해시 비교 방식)...")

        for i in range(1, 501): # 최대 500페이지
            
            print(f"[DEBUG] PDF Page-{i} 캡처 시도...")
            try:
                screenshot = capture_active_window(hwnd)
            except Exception as capture_err:
                print(f"[WARN] 캡처 실패(오류: {capture_err}). 루프를 중단합니다.")
                break
            
            # 1. 캡처한 이미지를 파일로 "먼저 저장"
            output_file_path = os.path.join(output_path, f"{base_filename}_page_{i:03}.png")
            screenshot.save(output_file_path, "PNG")
            
            # 2. 방금 저장된 파일의 해시 계산
            current_file_hash = _get_file_hash(output_file_path)
            
            print(f"[DEBUG] Page-{i} 비교: PrevHash={prev_file_hash}, CurrHash={current_file_hash}")

            # 3. 이전 파일 해시와 현재 파일 해시 비교
            if i > 1 and prev_file_hash == current_file_hash:
                # 두 파일 해시가 동일하면, PageDown이 안 먹힌 것 (문서 끝)
                print(f"[DEBUG] Page-{i}가 이전 페이지와 파일 해시가 동일하여 캡처를 중지합니다 (문서 끝).")
                
                # [사용자 요청] 마지막으로 저장된 중복 파일(page_i) 삭제
                try:
                    os.remove(output_file_path)
                    print(f"[DEBUG] 중복 저장된 {output_file_path} 파일을 삭제했습니다.")
                except Exception as e:
                    print(f"[WARN] 중복 파일 삭제 실패: {e}")
                
                break # 루프 중단
            
            # 4. (저장 성공) 현재 해시를 '이전 해시'로 저장하고 카운트 증가
            prev_file_hash = current_file_hash
            page_count += 1
            print(f"[DEBUG] PDF Page-{i} 캡처 및 저장 완료.")
            
            # 5. PageDown 키 전송
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(0.1) # 포커스 이동 대기
            print(f"[DEBUG] PageDown 키 전송 (pynput 방식)...")
            keyboard.press(Key.page_down)
            keyboard.release(Key.page_down)
            time.sleep(2.0) # 페이지 렌더링 대기 (넉넉하게 2초)
            
        print("[DEBUG] 캡처 완료. 뷰어 창에 WM_CLOSE 메시지 전송...")
        win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
        time.sleep(3.0) 

        print(f"[OK] PDF 문서 {page_count}페이지 이미지 저장 완료: {output_path}")

    except Exception as e:
        print(f"\n[!!!] PDF 변환 작업 중 심각한 오류 발생: {e}\n")
        raise RuntimeError(f"PDF 변환 중 오류 발생: {e}")
    
    return f"PDF 문서 {page_count}페이지 이미지를 저장 완료!\n{output_path}"


# --- 3. Convert To PDF (변경 없음) ---

def _numeric_sort_key(f):
    basename = os.path.splitext(os.path.basename(f))[0]
    try:
        # 파일명이 "slide_001.png" 같은 경우, "001"을 숫자로 변환하여 정렬
        # 숫자가 아닌 경우(예: "__MACOSX")는 basename으로 정렬
        return int(basename)
    except ValueError:
        return basename


def convert_to_pdf(target_dir, output_file):
    """
    지정된 디렉터리 내의 이미지 파일들을 모아 하나의 PDF 파일로 변환
    (이전 버전의 ZIP 파일 처리 로직 제거됨)
    """
    
    # 1. 이미지 파일 확장자 정의
    img_extensions = ('.png', '.jpg', '.jpeg', '.bmp', '.gif')
    
    # 2. 지정된 디렉터리에서 이미지 파일 목록을 가져옵니다.
    # glob.glob을 사용하여 모든 파일을 검색하고, 확장자를 확인하여 필터링합니다.
    target_dir = os.path.abspath(target_dir)
    output_file = os.path.abspath(output_file)

    image_files = [f for f in glob.glob(os.path.join(target_dir, "*")) 
                   if os.path.splitext(f)[1].lower() in img_extensions]
                   
    if not image_files:
        raise Exception(f"'{target_dir}' 디렉터리 내에 변환할 수 있는 이미지 파일이 없습니다. (지원 확장자: {img_extensions})")
        
    # 3. 파일 목록을 순서대로 정렬 (slide_001.png, slide_002.png 순서 보장)
    image_files.sort(key=_numeric_sort_key)
    
    # 4. Pillow Image 객체로 로드 (PDF 변환을 위해 RGB로 변환)
    # PIL.Image.open() 시 파일이 잠기는 것을 방지하기 위해 .convert('RGB')까지 처리
    try:
        images_pil = [Image.open(f).convert('RGB') for f in image_files]
    except Exception as e:
        raise RuntimeError(f"이미지 파일을 로드하는 중 오류 발생: {e}")

    
    # 5. PDF 파일 저장 경로 설정 (output_file은 app_window.py에서 이미 전체 경로를 받음)
    pdf_path = output_file
    
    # 6. 첫 번째 이미지를 기준으로 PDF를 생성하고 나머지 이미지들을 추가합니다.
    if images_pil:
        images_pil[0].save(
            pdf_path,
            save_all=True,
            append_images=images_pil[1:]
        )
    else:
        # 이 else 블록은 2단계에서 이미 처리되었으나, 안전을 위해 남겨둡니다.
        raise Exception("변환할 이미지가 준비되지 않았습니다.")
    
    
    # 이전 버전에서 사용되던 shutil, tempfile 관련 로직은 제거되었습니다.
    
    return f"PDF 변환 완료!\n총 {len(image_files)}개의 이미지를 {pdf_path}로 병합했습니다."
